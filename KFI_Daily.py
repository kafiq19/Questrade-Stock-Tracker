'''
This script is designed to automate the workflow of 
tracking stocks currently owned, sold and watching. 
Takes advantage of the Questrade API for which an
access code must be derived through them.

Author: Khaleel Arfeen, k.a@unb.ca 
Date: May 05, 2018
Version: 1.0.1 -- basic tracking 
'''

import os
import sys
import datetime
import qt_api_interact as qt_api
import styles
from openpyxl import load_workbook

################
# User Variables
file_name = "!MyTrackers_Sample.xlsx" #change this to your excel file, keep in same directory as this file
worksheet_name = "Currents" #enter worksheet name if needed, leave blank if one sheet
reporting_currency = "CAD"
refresh_token = "3kyJ-t_xAcsA14JXYN9ciIb-UWY-F9py0"
num_accounts = 10 #if you have >10 accounts in excel, change this to the correct number
btc_holdings = 0.75623876
non_qt_accounts = {"(EJ)": 0, "BTC": qt_api.get_btc_price()*btc_holdings}

# Global Variables
access_token = ""
uri = ""
account_names = {} #key is number, value is {acct#: total_equity..}
all_positions = {} #key is ticker, value is current price {aaple: 188.84..}
################

# Load in the excel file
try:
	wb = load_workbook(file_name, data_only=True)
	print "Sucessfully loaded {}".format(file_name)
except IOError as e:
	sys.exit('ERROR: Could not open the excel file {}'.format(file_name))

# Load Sheet
# output current sheet sheet
def load_worksheet():
	if worksheet_name == "":
		sheet = wb.active
		print "loaded default worksheet"
		return sheet
	else:
		sheet_names = wb.sheetnames # Get sheet names
		if worksheet_name in sheet_names:
			sheet = wb[worksheet_name]
			print "loaded {} worksheet".format(worksheet_name)
			return sheet
		else:
			print "error loading worksheet name. Set in user variables or keep blank for default"
			print "will try to load default sheet.."
			try:
				sheet = wb.active
			except IOError as e:
				sys.exit('ERROR: Could not open the excel worksheet {}'.format(worksheet_name))

# Prep Worksheet
# input worksheet, output worksheet
def prep_worksheet(sheet):
	sheet.insert_cols(3) #insert new column
	now = datetime.datetime.now()
	sheet['C1'] = now.strftime("%d-%B-%Y %H:%M") #add date header
	return sheet

#Obtain Questrade Access token and URL
def get_access_qt(refresh):
	global access_token, uri
	access_token, uri = qt_api.get_access_code(refresh)

#Establish Connection to the Questrade Server
#Return account information reply from QT
def connect_questrade():
	try:
		account_info = qt_api.establish_connection(access_token, uri)
		reply_message = account_info.get('message')
	except IOError as e:
		sys.exit('ERROR: Did not receive a reply from the Questrade server')

	if reply_message == "Access token is invalid":
		sys.exit('ERROR: Check the access key and URL')
	else:
		return account_info


#get the questrade account types and number and set globally
#iput is output of connect_questrade()
def fetch_accounts(account_info):
	#access only the account information dict
	account_info = account_info.get('accounts')

	#get the account type and number from the response as tuple-list
	for account in account_info:
		account_names[int(account.get("number"))] = [account.get("type")]

#returns the total_equity balance of the account
#input global access_token, uri and account_names(dict)
def fetch_account_values(access_code, url, account):
	for key, value in account.iteritems():
		account_values = qt_api.get_account_values(access_code, url, key) #inputs account #
		account_values = account_values.get("combinedBalances") #filter combinedBalances from response

		#filter canadian balances
		for currency_value in account_values:
			if currency_value.get("currency") == reporting_currency:
				account_values = currency_value
			else:
				pass
		account_values = account_values.get('totalEquity')
		value.append(round(account_values, 2))
		account_names[key] = value

#sets global variable all_positions as {ticker: currentPrice,..}
def fetch_account_positions(access_code, url, account):
	for key, value in account.iteritems():
		positions = qt_api.get_account_positions(access_code, url, value[0])

		for equity in positions.get('positions'):
			all_positions[equity.get('symbol')] = equity.get('currentPrice')

# fills the account values section of the xl sheet
def add_account_info_to_xl(sheet):
	for row in sheet.iter_rows('A1:A{}'.format(num_accounts)):
		
		#add total amounts
		current_row = row[0].row
		if row[0].value == "Total":
			sheet['C{}'.format(current_row)] = "=SUM(C1:C{})".format(current_row-1)
			sheet['C{}'.format(current_row+1)] = \
				"=((C{}-$B${})/$B${})*100".format(current_row, current_row, current_row)
			sheet['C{}'.format(current_row-1)].border = styles.double_bottom_border
			return sheet
		
		#add values to accounts in account names
		elif row[0].value in account_names:
			account_number = row[0].value
			account_value = account_names.get(account_number)[1]
			sheet['C{}'.format(current_row)] = account_value

		#add values to non-questrade accounts in non_qt_accounts
		elif row[0].value in non_qt_accounts:
			account_value = non_qt_accounts.get(row[0].value)
			sheet['C{}'.format(current_row)] = account_value

		else:
			print "Adding account info.."

# fills in the equities section of the xl sheet
def add_equity_info_to_xl(sheet):
	begin_index = 0
	end_index = 0
	switch = 1
	
	# get start and end of equity section
	for cell in sheet['A']:
		if cell.value == "EQUITIES":
			begin_index = cell.row
		elif cell.value == "EQUITIES END":
			end_index = cell.row
		else:
			pass

	for cell in sheet.iter_rows('B{}:B{}'.format(begin_index+1, end_index-1)):
		current_row = cell[0].row

		# set ticker price
		if switch == 1:
			if cell[0].value in all_positions:
				sheet['C{}'.format(current_row)] = all_positions.get(cell[0].value)
				switch = 2
			elif cell[0].value in non_qt_accounts:
				sheet['C{}'.format(current_row)] = non_qt_accounts.get(cell[0].value)
				switch = 2
			else:
				sheet['C{}'.format(current_row)] = 0
				switch = 2

		# set overall %gain (current$-initial$)/intitial$
		elif switch == 2:
			pass
			sheet['C{}'.format(current_row)] = \
				"=((C{}-$B{})/$B{})*100".format(current_row-1, current_row+1, current_row+1)          
			switch = 3

		# set days % gain 
		else:
			sheet['C{}'.format(current_row)] = \
				"=((C{}-$D{})/$D{})*100".format(current_row-2, current_row-2, current_row-2)
			sheet['C{}'.format(current_row)].border = styles.thin_bottom_border
			switch = 1

# __Main__:
excel_sheet = load_worksheet()
excel_sheet = prep_worksheet(excel_sheet)
get_access_qt(refresh_token)
account_info = connect_questrade()
fetch_accounts(account_info)
fetch_account_values(access_token, uri, account_names)
fetch_account_positions(access_token, uri, account_names)
excel_sheet = add_account_info_to_xl(excel_sheet)
print "Adding equity section.."
excel_sheet = add_equity_info_to_xl(excel_sheet)
print "Saving output.."
os.remove(file_name)
wb.save(file_name)
print "Complete"

'''
BUGS:
- insert_cols(3) giving call not recognized on mac
- some colomns for %ROI are being deleted
'''
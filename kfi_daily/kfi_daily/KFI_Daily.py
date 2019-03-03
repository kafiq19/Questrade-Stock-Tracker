'''
This script is designed to automate the workflow of 
tracking stocks currently owned, sold and watching. 
Takes advantage of the Questrade API for which an
access code must be derived through them.
 
Author: Khaleel Arfeen, k.a@unb.ca 
Date: May 05, 2018
Version: 1.1.1 -- class implementation
'''
 
import os
import sys
import datetime
import qt_api_interact as qt_api
import styles
from openpyxl import load_workbook

class KFI_Daily:

    def __init__(self):
        self.xl_file = "MyTracker.xlsx" #change this to your excel file, keep in same directory as this file 
        self.refresh_token = "AOb_5NnC4s8gMezxiS55mtKTsmbTk1950" #update from questrade if first time or expired
        self.worksheet_name = 'Currents' #enter worksheet name if needed, leave blank if one sheet
        self.reporting_currency = "CAD"
        self.num_accounts = 10 #if you have >10 accounts(not equities!) in excel, change this to the correct number
        self.btc_holdings = 0
        self.non_qt_accounts = {"(EJ)": 0, "BTC": qt_api.get_btc_price()*self.btc_holdings, "other": 0}

        self.access_token = ""
        self.uri = ""
        self.account_info = False
        self.account_names = {} #key is number, value is {acct#: total_equity..}
        self.all_positions = {} #key is ticker, value is current price {aaple: 188.84..}
        self.wb_handle = False
        self.sheet_handle = False

    # Load in the excel file
    def load_file(self):
	    try:
	        self.wb_handle = load_workbook(self.xl_file, data_only=True)
	        print "Sucessfully loaded {}".format(self.xl_file)
	    except IOError as e:
	        sys.exit('ERROR: Could not open the excel file {}'.format(self.xl_file))
 
    # Load Sheet
    # output current sheet sheet
    def load_worksheet(self):
        if self.worksheet_name == "":
            self.sheet_handle = self.wb_handle.active
            print "loaded default worksheet"
        else:
            sheet_names = self.wb_handle.sheetnames # Get sheet names
            if self.worksheet_name in sheet_names:
                self.sheet_handle = self.wb_handle[self.worksheet_name]
                print "loaded {} worksheet".format(self.worksheet_name)
            else:
                print "error loading worksheet name. Set in user variables or keep blank for default"
                print "will try to load default sheet.."
                try:
                    self.sheet_handle = self.wb_handle.active
                except IOError as e:
                    sys.exit('ERROR: Could not open the excel worksheet {}'.format(self.sheet_handle))
 
    # Prep Worksheet
    # input worksheet, output worksheet
    def prep_worksheet(self):
        sheet = self.sheet_handle
        sheet.insert_cols(3) #insert new column
        now = datetime.datetime.now()
        sheet['C1'] = now.strftime("%d-%B-%Y %H:%M") #add date header
 
    #Obtain Questrade Access token and URL
    def get_access_qt(self):
        self.access_token, self.uri = qt_api.get_access_code(self.refresh_token) 
 
	#Establish Connection to the Questrade Server
	#Return account information reply from QT
    def connect_questrade(self):
        try:
            account_info = qt_api.establish_connection(self.access_token, self.uri)
            reply_message = account_info.get('message')
        except IOError as e:
            sys.exit('ERROR: Did not receive a reply from the Questrade server - Try Creating New Refresh Code')
	 
        if reply_message == "Access token is invalid":
            sys.exit('ERROR: Access Not Granted - Try Creating New Refresh Code')
        else:
            self.account_info = account_info
 
 
	#get the questrade account types and number and set globally
    def fetch_accounts(self):
        #access only the account information dict
        account_info = self.account_info.get('accounts')
	 
        #get the account type and number from the response as tuple-list
        for account in account_info:
            self.account_names[int(account.get("number"))] = [account.get("type")]
 
    #returns the total_equity balance of the account
    #input global access_token, uri and account_names(dict)
    def fetch_account_values(self):
        for key, value in self.account_names.iteritems():
            account_values = qt_api.get_account_values(self.access_token, self.uri, key) #inputs account #
            account_values = account_values.get("combinedBalances") #filter combinedBalances from response

            #filter canadian balances
            for currency_value in account_values:
                if currency_value.get("currency") == self.reporting_currency:
                    account_values = currency_value
                    account_values = account_values.get('totalEquity')
                    value.append(round(account_values, 2))
                else:
                    pass
            self.account_names[key] = value
 
    #sets global variable all_positions as {ticker: currentPrice,..}
    def fetch_account_positions(self):
        for key, value in self.account_names.iteritems():
            positions = qt_api.get_account_positions(self.access_token, self.uri, key)
     
            for equity in positions.get('positions'):
                self.all_positions[equity.get('symbol')] = \
                    [equity.get('openQuantity'), equity.get('averageEntryPrice'), \
                    equity.get('currentPrice'), equity.get('openPnl'), equity.get('totalCost')]

        print "Successfully fetched account information"

 
    # fills the account values section of the xl sheet
    def add_account_info_to_xl(self):
        sheet = self.sheet_handle
        for row in sheet['A']:

            #add total amounts
            current_row = row.row
            if row.value == "Total":
                sheet['C{}'.format(current_row)] = "=SUM(C1:C{})".format(current_row-1)
                sheet['C{}'.format(current_row+1)] = \
                    "=((C{}-$B${})/$B${})*100".format(current_row, current_row, current_row)
                sheet['C{}'.format(current_row-1)].border = styles.double_bottom_border

                self.sheet_handle = sheet
                return
         
            #add values to accounts in account names
            elif row.value in self.account_names:
                account_number = row.value
                account_value = self.account_names.get(account_number)[1]
                sheet['C{}'.format(current_row)] = account_value
 
            #add values to non-questrade accounts in non_qt_accounts
            elif row.value in self.non_qt_accounts:
                account_value = self.non_qt_accounts.get(row.value)
                sheet['C{}'.format(current_row)] = account_value
 
            else:
                pass
                
        print "Successfully added account info"

    # fills in the equities section of the xl sheet
    def add_equity_info_to_xl(self):
        sheet = self.sheet_handle

        begin_index = 0
        switch = 0
         
        # get start and end of equity section
        for cell in sheet['A']:
            if cell.value == "EQUITIES":
                begin_index = cell.row +1
            else:
                pass
        
        sorted_equities = sorted(self.all_positions.keys())

        for equity in sorted_equities:
            equity_data = self.all_positions.get(equity)
            sheet['C{}'.format(begin_index)] = equity #set name
            sheet['C{}'.format(begin_index)].font = styles.bold_font
            sheet['C{}'.format(begin_index+1)] = equity_data[0] #set number of shares
            sheet['C{}'.format(begin_index+2)] = equity_data[1] #set buy price
            sheet['C{}'.format(begin_index+3)] = equity_data[2] #set current price
            sheet['C{}'.format(begin_index+4)] = equity_data[3] #set $roi
            sheet['C{}'.format(begin_index+5)] = (equity_data[4]/equity_data[4])*100 #set %roi
            begin_index = begin_index+6

        print "Successfully added equity info"
        print "Saving output.."
 

#__MAIN__:
tracker = KFI_Daily()
tracker.load_file()
tracker.load_worksheet()
tracker.prep_worksheet()
tracker.get_access_qt()
tracker.connect_questrade()
tracker.fetch_accounts()
tracker.fetch_account_values()
tracker.fetch_account_positions()
tracker.add_account_info_to_xl()
tracker.add_equity_info_to_xl()
tracker.wb_handle.save(tracker.xl_file)
print "Complete"
#import pdb; pdb.set_trace()


'''
BUGS:
- insert_cols(3) giving call not recognized on mac
- some colomns for %ROI are being deleted

Wants:
- web interface
- real time tracking: impact on balance, analytics, trading
- news: company, market, technical analysis
- one place for everything
- automatic run for history
- one click interface, easy user interface
'''